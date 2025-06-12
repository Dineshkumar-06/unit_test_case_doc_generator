import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import os

import copy
from openpyxl.cell.cell import Cell  # Import Cell class to check type


def is_row_empty(row):
    return all(cell.value is None for cell in row)

# def get_last_data_column(sheet):
#     max_col = 1
#     for row in sheet.iter_rows():
#         for i, cell in enumerate(row):
#             if cell.value:
#                 max_col = max(max_col, i + 1)
#     return max_col

def get_last_data_column(sheet):
    max_col = 1
    for row in sheet.iter_rows(min_row=1, max_row=10):  # Check only top 10 rows
        for i, cell in enumerate(row, 1):
            if cell.value not in (None, ""):
                max_col = max(max_col, i)
    return max_col

def get_last_data_row(sheet):
    for row in reversed(list(sheet.iter_rows())):
        if not is_row_empty(row):
            return row[0].row
    return 1

# def add_testcase_columns(sheet):
#     last_col = get_last_data_column(sheet)
#     last_row = get_last_data_row(sheet)

#     headers = ["EXPECTED OUTPUT", "DEFECT SEVERITY", "REMARKS"]
#     fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#     align = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     for i, header in enumerate(headers):
#         col_idx = last_col + i + 1
#         col_letter = get_column_letter(col_idx)
#         sheet[f"{col_letter}1"] = header
#         sheet[f"{col_letter}1"].alignment = align

#         if header == "EXPECTED OUTPUT":
#             for row in range(2, last_row + 1):
#                 if all(sheet.cell(row=row, column=col).value in [None, ""] for col in range(1, 5)):
#                     continue
#                 cell = sheet[f"{col_letter}{row}"]
#                 cell.value = "Pass"
#                 cell.fill = fill
#                 cell.alignment = align

def add_testcase_columns(sheet):
    last_col = get_last_data_column(sheet)
    # last_row = sheet.max_row
    last_row = get_last_data_row(sheet)

    headers = ["EXPECTED OUTPUT", "DEFECT SEVERITY", "REMARKS"]
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    column_widths = {
        "EXPECTED OUTPUT": 20,
        "DEFECT SEVERITY": 18,
        "REMARKS": 25
    }

    for i, header in enumerate(headers):
        col_idx = last_col + i + 1
        col_letter = get_column_letter(col_idx)

        # Unmerge if necessary before assigning new headers
        for rng in sheet.merged_cells.ranges:
            if f"{col_letter}1" in str(rng):
                sheet.unmerge_cells(str(rng))
                break

        sheet[f"{col_letter}1"] = header
        sheet[f"{col_letter}1"].alignment = align

        sheet.column_dimensions[col_letter].width = column_widths.get(header, 15)

        if header == "EXPECTED OUTPUT":
            for row in range(2, last_row + 1):
                if all(sheet.cell(row=row, column=col).value in [None, ""] for col in range(1, last_col + 1)):
                    continue
                cell = sheet[f"{col_letter}{row}"]
                cell.value = "Pass"
                cell.fill = fill
                cell.alignment = align


def delete_static_sheets(wb):
    for name in ["ReadMe", "Title"]:
        if name in wb.sheetnames:
            std = wb[name]
            wb.remove(std)

def insert_static_sheets(wb_target, static_sheet_file, flag):
    wb_static = openpyxl.load_workbook(static_sheet_file)

    for sheet in reversed(wb_static.sheetnames):
        source = wb_static[sheet]
        if flag == 0:
            new_sheet = wb_target.create_sheet(title=sheet, index=0)
        else:
            new_sheet = wb_target.create_sheet(title=sheet)


        for row in source.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell):
                    continue  # Skip MergedCell or other non-standard cells

                # new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                if not isinstance(new_cell, MergedCell):
                    new_cell.value = cell.value

                if cell.has_style:
                    if cell.font:
                        new_cell.font = copy.copy(cell.font)
                    if cell.border:
                        new_cell.border = copy.copy(cell.border)
                    if cell.fill:
                        new_cell.fill = copy.copy(cell.fill)
                    if cell.number_format:
                        new_cell.number_format = copy.copy(cell.number_format)
                    if cell.protection:
                        new_cell.protection = copy.copy(cell.protection)
                    if cell.alignment:
                        new_cell.alignment = copy.copy(cell.alignment)

        # Copy merged cell ranges
        for merged_range in source.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))

        # Copy column widths
        for col in source.column_dimensions:
            new_sheet.column_dimensions[col].width = source.column_dimensions[col].width

        # Copy row heights
        for row_idx in source.row_dimensions:
            new_sheet.row_dimensions[row_idx].height = source.row_dimensions[row_idx].height

def add_testcase_columns_func(sheet, start_row, end_row, start_col):
    """
    Adds 'Pass' in columns starting at start_col (Expected Output, etc.) 
    for rows from start_row to end_row.
    """
    EXPECTED_OUTPUT_COL = start_col
    DEFECT_SEVERITY_COL = EXPECTED_OUTPUT_COL + 1
    REMARKS_COL = DEFECT_SEVERITY_COL + 1

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Add headers
    sheet.cell(row=start_row - 1, column=EXPECTED_OUTPUT_COL, value="EXPECTED OUTPUT")
    sheet.cell(row=start_row - 1, column=DEFECT_SEVERITY_COL, value="DEFECT SEVERITY")
    sheet.cell(row=start_row - 1, column=REMARKS_COL, value="REMARKS")

    for row in range(start_row, end_row + 1):
        if all(sheet.cell(row=row, column=col).value in [None, ""] for col in range(1, 5)):
            continue

        sheet.cell(row=row, column=EXPECTED_OUTPUT_COL, value="Pass").fill = green_fill

def append_age_workbook_to_single_sheet(wb_target, age_file_path, add_testcase_columns_func):
    import copy
    import openpyxl
    from openpyxl.styles import PatternFill

    age_wb = openpyxl.load_workbook(age_file_path)
    combined_sheet = wb_target.create_sheet("Age")

    current_row = 1
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for sheet in age_wb.worksheets:
        merge_ranges = list(sheet.merged_cells.ranges)
        merged_bounds = [r.bounds for r in merge_ranges]

        # Copy cell values and styles
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                target_row = current_row + row_idx - 1
                target_col = cell.column

                # Check if this cell is *not* part of a to-be-merged range (except top-left)
                in_merged = False
                for (min_col, min_row, max_col, max_row) in merged_bounds:
                    if min_row == cell.row and min_col == cell.column:
                        break  # Top-left cell of merged range
                    if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                        in_merged = True
                        break

                if in_merged:
                    continue  # Skip writing to merged region inner cells

                # target_cell = combined_sheet.cell(
                #     row=target_row,
                #     column=target_col,
                #     value=cell.value
                # )

                target_cell = combined_sheet.cell(row=target_row, column=target_col)
                if not isinstance(target_cell, MergedCell):
                    target_cell.value = cell.value

                if cell.has_style:
                    target_cell.font = copy.copy(cell.font)
                    target_cell.border = copy.copy(cell.border)
                    target_cell.fill = copy.copy(cell.fill)
                    target_cell.number_format = copy.copy(cell.number_format)
                    target_cell.protection = copy.copy(cell.protection)
                    target_cell.alignment = copy.copy(cell.alignment)

        # Copy merged cells
        for merge_range in merge_ranges:
            min_col, min_row, max_col, max_row = merge_range.bounds
            combined_sheet.merge_cells(
                start_row=current_row + min_row - 1,
                start_column=min_col,
                end_row=current_row + max_row - 1,
                end_column=max_col
            )

        current_row += sheet.max_row + 1  # Space between blocks

        # Add yellow separator row
        YELLOW_ROW_LENGTH = 50
        for col_idx in range(1, YELLOW_ROW_LENGTH + 1):
            combined_sheet.cell(row=current_row, column=col_idx).fill = yellow_fill
        current_row += 2

    # Apply test case columns blockwise
    row = 1
    while row <= combined_sheet.max_row:
        row_values = []
        for col in range(1, combined_sheet.max_column + 1):
            cell = combined_sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                row_values.append("")  # Treat merged (non-top-left) as empty
            else:
                value = str(cell.value).strip().lower() if cell.value else ""
                row_values.append(value)

        if "general" in row_values or "open" in row_values:
            # Determine where "relaxation years" is in the row
            try:
                start_col = row_values.index("relaxation years") + 2  # next column
            except ValueError:
                start_col = sheet.max_column + 1

            start_row = row + 1
            end_row = start_row

            while end_row <= combined_sheet.max_row:
                if all(combined_sheet.cell(row=end_row, column=col).value in [None, ""] for col in range(1, 6)):
                    break
                end_row += 1

            add_testcase_columns_func(combined_sheet, start_row - 1, end_row - 1, start_col)

            row = end_row
        else:
            row += 1

def append_testcase_other_details(sheet):
    last_col = get_last_data_column(sheet)
    # last_row = get_last_data_row(sheet)
    
    headers = ["EXPECTED OUTPUT", "DEFECT SEVERITY", "REMARKS"]
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    column_widths = {
        "EXPECTED OUTPUT": 20,
        "DEFECT SEVERITY": 18,
        "REMARKS": 25
    }

    for i, header in enumerate(headers):
        col_idx = last_col + i + 1
        col_letter = get_column_letter(col_idx)
        sheet[f"{col_letter}1"] = header
        sheet[f"{col_letter}1"].alignment = align
        sheet.column_dimensions[col_letter].width = column_widths.get(header, 15)

        if header == "EXPECTED OUTPUT":
            row = 2
            cell = sheet[f"{col_letter}{row}"]
            cell.value = "Pass"
            cell.fill = fill
            cell.alignment = align

def process_excel(input_path, static_sheet_file, age_sheet_file, eligible_sheet_file, additional_docs):
    wb = openpyxl.load_workbook(input_path)

    delete_static_sheets(wb)
    insert_static_sheets(wb, static_sheet_file, 0)
    print("static sheets inserted!")

    append_age_workbook_to_single_sheet(wb, age_sheet_file, add_testcase_columns_func)
    print("age sheets inserted!")
    insert_static_sheets(wb, eligible_sheet_file, 1)
    print("eligiblility sheets inserted!")

    # if additional_docs != []:
    #     for i in additional_docs:
    #         insert_static_sheets(wb, i, 1)

    if additional_docs != []:
        print("additional sheets found!")
        print(additional_docs)
        for i in additional_docs:
            try:
                insert_static_sheets(wb, i, 1)
                print(f"{additional_docs} {i} sheets inserted!")
            except Exception as e:
                messagebox.showerror("Error", f"Error inserting dynamic file '{i}':\n{str(e)}")

    for sheet in wb.worksheets:
        if sheet.title not in ["ReadMe", "Title","Age", "Other Details"]:
            add_testcase_columns(sheet)
            print("testcase column inserted for all!")
        elif sheet.title == "Other Details":
            append_testcase_other_details(sheet)
            print("testcase column inserted for other details!")

    print("testcase document done - step before saving the file!")
    base, ext = os.path.splitext(input_path)
    output_path = f"{base}_testcases{ext}"
    print(output_path)
    wb.save(output_path)
    print("Success!")
    return output_path

# ---- Tkinter GUI ----
def browse_input():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        input_path.set(filepath)

def browse_static():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        static_path.set(filepath)

def browse_age():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        age_path.set(filepath)

def browse_eligible():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        eligible_path.set(filepath)

# ----------------------------------------------------------------------------------------------

# Dynamic fields functions

def update_fields():
    selected = radio_var.get()

    # Hide all dynamic widgets first
    for widget in dynamic_frame.winfo_children():
        widget.pack_forget()

    if selected == "Y":
        # tk.Label(dynamic_frame, text="Field for Option 1").pack()
        # tk.Entry(dynamic_frame).pack()
        button_frame.pack(pady=5)
        dynamic_frame.pack(pady=10)
        
    elif selected == "N":
        # button_frame.pack_forget()
        dynamic_frame.pack_forget()
        

def add_fields():
    var = tk.StringVar()

    field_frame = tk.Frame(dynamic_frame)
    field_frame.pack(pady=5, fill='x')

    def browse():
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            var.set(filepath)

    tk.Label(field_frame, text="Additional File:").pack(pady=5)
    tk.Entry(field_frame,textvariable=var, width=50).pack()
    tk.Button(field_frame, text="Browse", command=browse).pack(pady=5)

    dynamic_field_frames.append(field_frame)
    additional_docs.append(var)

def remove_fields():
    if dynamic_field_frames:
        last_frame = dynamic_field_frames.pop() # Get the last added frame
        additional_docs.pop() 
        last_frame.destroy() 
        scrollable_frame.update_idletasks()

# ----------------------------------------------------------------------------------------------

def run_processing():
    dynamic_docs = [var.get() for var in additional_docs]
    if dynamic_docs != []:
        print(dynamic_docs)
    else:
        print("No dynamic docs found!")

    if not input_path.get() or not static_path.get() or not age_path.get() or not eligible_path.get():
        messagebox.showerror("Missing Input", "Please select all the files.")
        return
    try:
        output = process_excel(input_path.get(), static_path.get(), age_path.get(), eligible_path.get(), dynamic_docs)
        messagebox.showinfo("Success", f"Test Case Document generated successfully!\n\nTest Case Document created at:\n{output}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n{str(e)}")

    # Reset fields
    input_path.set("")
    static_path.set("")
    age_path.set("")
    eligible_path.set("")
    radio_var.set("N")
    dynamic_frame.pack_forget()

# Main frame
app = tk.Tk()
app.title("Unit Testcase Document Generator")
app.geometry("1000x1000")

input_path = tk.StringVar()
static_path = tk.StringVar()
age_path = tk.StringVar()
eligible_path = tk.StringVar()
radio_var = tk.StringVar(value="N")

dynamic_field_frames = []
additional_docs = []

canvas = tk.Canvas(app)
scrollbar = tk.Scrollbar(app, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

tk.Label(scrollable_frame, text="Input SOW Excel File (SOW.xlsx)").pack(pady=5)
tk.Entry(scrollable_frame, textvariable=input_path, width=50).pack()
tk.Button(scrollable_frame, text="Browse", command=browse_input).pack(pady=5)

tk.Label(scrollable_frame, text="Static Sheets File (static_sheets.xlsx)").pack(pady=5)
tk.Entry(scrollable_frame, textvariable=static_path, width=50).pack()
tk.Button(scrollable_frame, text="Browse", command=browse_static).pack(pady=5)

tk.Label(scrollable_frame, text="Age Sheet (age.xlsx)").pack(pady=5)
tk.Entry(scrollable_frame, textvariable=age_path, width=50).pack()
tk.Button(scrollable_frame, text="Browse", command=browse_age).pack(pady=5)

tk.Label(scrollable_frame, text="Eligibility Sheet (eligibility.xlsx)").pack(pady=5)
tk.Entry(scrollable_frame, textvariable=eligible_path, width=50).pack()
tk.Button(scrollable_frame, text="Browse", command=browse_eligible).pack(pady=5)

# tk.Label(app, text="Is there any additonal files?").pack( padx=5, pady=2)
# tk.Radiobutton(app, text="Yes", variable=radio_var, value="Y", command=update_fields).pack( padx=5, pady=2)
# tk.Radiobutton(app, text="No", variable=radio_var, value="N", command=update_fields).pack( padx=5, pady=2)

radio_frame = tk.Frame(scrollable_frame)
radio_frame.pack(padx=350, pady=15, anchor="w")  # Align left

tk.Label(radio_frame, text="Is there any additional files?").pack(side="left", padx=5)
tk.Radiobutton(radio_frame, text="Yes", variable=radio_var, value="Y", command=update_fields).pack(side="left", padx=10)
tk.Radiobutton(radio_frame, text="No", variable=radio_var, value="N", command=update_fields).pack(side="left", padx=10)

# Frame to hold dynamic fields
dynamic_frame = tk.Frame(scrollable_frame)
button_frame = tk.Frame(dynamic_frame)
dynamic_frame.pack(pady=5, fill="x", anchor="w")


tk.Button(button_frame, text="Add File input", command=add_fields).pack()
tk.Button(button_frame, text="Remove Last File input", command=remove_fields).pack()

tk.Button(scrollable_frame, text="Generate Test Case Document", command=run_processing, bg="green", fg="white").pack(pady=20)

update_fields()

app.mainloop()

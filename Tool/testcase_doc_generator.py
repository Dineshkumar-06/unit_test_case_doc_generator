import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import os

import copy
from openpyxl.cell.cell import Cell  # Import Cell class to check type


def is_row_empty(row):
    return all(cell.value is None for cell in row)

def get_last_data_column(sheet):
    max_col = 1
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value:
                max_col = max(max_col, i + 1)
    return max_col

def get_last_data_row(sheet):
    for row in reversed(list(sheet.iter_rows())):
        if not is_row_empty(row):
            return row[0].row
    return 1

def add_testcase_columns(sheet):
    last_col = get_last_data_column(sheet)
    last_row = get_last_data_row(sheet)

    headers = ["EXPECTED OUTPUT", "DEFECT SEVERITY", "REMARKS"]
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, header in enumerate(headers):
        col_idx = last_col + i + 1
        col_letter = get_column_letter(col_idx)
        sheet[f"{col_letter}1"] = header
        sheet[f"{col_letter}1"].alignment = align

        if header == "EXPECTED OUTPUT":
            for row in range(2, last_row + 1):
                if all(sheet.cell(row=row, column=col).value in [None, ""] for col in range(1, 5)):
                    continue
                cell = sheet[f"{col_letter}{row}"]
                cell.value = "Pass"
                cell.fill = fill
                cell.alignment = align

def delete_static_sheets(wb):
    for name in ["ReadMe", "Title"]:
        if name in wb.sheetnames:
            std = wb[name]
            wb.remove(std)

def insert_static_sheets(wb_target, static_sheet_file, flag):
    wb_static = openpyxl.load_workbook(static_sheet_file)

    for sheet in reversed(wb_static.sheetnames):
        source = wb_static[sheet]
        if flag == 0:
            new_sheet = wb_target.create_sheet(title=sheet, index=0)
        else:
            new_sheet = wb_target.create_sheet(title=sheet)


        for row in source.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell):
                    continue  # Skip MergedCell or other non-standard cells

                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                if cell.has_style:
                    if cell.font:
                        new_cell.font = copy.copy(cell.font)
                    if cell.border:
                        new_cell.border = copy.copy(cell.border)
                    if cell.fill:
                        new_cell.fill = copy.copy(cell.fill)
                    if cell.number_format:
                        new_cell.number_format = copy.copy(cell.number_format)
                    if cell.protection:
                        new_cell.protection = copy.copy(cell.protection)
                    if cell.alignment:
                        new_cell.alignment = copy.copy(cell.alignment)

        # Copy merged cell ranges
        for merged_range in source.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))

        # Copy column widths
        for col in source.column_dimensions:
            new_sheet.column_dimensions[col].width = source.column_dimensions[col].width

        # Copy row heights
        for row_idx in source.row_dimensions:
            new_sheet.row_dimensions[row_idx].height = source.row_dimensions[row_idx].height

def add_testcase_columns_func(sheet, start_row, end_row, start_col):
    """
    Adds 'Pass' in columns starting at start_col (Expected Output, etc.) 
    for rows from start_row to end_row.
    """
    EXPECTED_OUTPUT_COL = start_col
    DEFECT_SEVERITY_COL = EXPECTED_OUTPUT_COL + 1
    REMARKS_COL = DEFECT_SEVERITY_COL + 1

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Add headers
    sheet.cell(row=start_row - 1, column=EXPECTED_OUTPUT_COL, value="EXPECTED OUTPUT")
    sheet.cell(row=start_row - 1, column=DEFECT_SEVERITY_COL, value="DEFECT SEVERITY")
    sheet.cell(row=start_row - 1, column=REMARKS_COL, value="REMARKS")

    for row in range(start_row, end_row + 1):
        if all(sheet.cell(row=row, column=col).value in [None, ""] for col in range(1, 5)):
            continue

        sheet.cell(row=row, column=EXPECTED_OUTPUT_COL, value="Pass").fill = green_fill

def append_age_workbook_to_single_sheet(wb_target, age_file_path, add_testcase_columns_func):
    import copy
    import openpyxl
    from openpyxl.styles import PatternFill

    age_wb = openpyxl.load_workbook(age_file_path)
    combined_sheet = wb_target.create_sheet("Age")

    current_row = 1
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for sheet in age_wb.worksheets:
        merge_ranges = list(sheet.merged_cells.ranges)
        merged_bounds = [r.bounds for r in merge_ranges]

        # Copy cell values and styles
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                target_row = current_row + row_idx - 1
                target_col = cell.column

                # Check if this cell is *not* part of a to-be-merged range (except top-left)
                in_merged = False
                for (min_col, min_row, max_col, max_row) in merged_bounds:
                    if min_row == cell.row and min_col == cell.column:
                        break  # Top-left cell of merged range
                    if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                        in_merged = True
                        break

                if in_merged:
                    continue  # Skip writing to merged region inner cells

                target_cell = combined_sheet.cell(
                    row=target_row,
                    column=target_col,
                    value=cell.value
                )

                if cell.has_style:
                    target_cell.font = copy.copy(cell.font)
                    target_cell.border = copy.copy(cell.border)
                    target_cell.fill = copy.copy(cell.fill)
                    target_cell.number_format = copy.copy(cell.number_format)
                    target_cell.protection = copy.copy(cell.protection)
                    target_cell.alignment = copy.copy(cell.alignment)

        # Copy merged cells
        for merge_range in merge_ranges:
            min_col, min_row, max_col, max_row = merge_range.bounds
            combined_sheet.merge_cells(
                start_row=current_row + min_row - 1,
                start_column=min_col,
                end_row=current_row + max_row - 1,
                end_column=max_col
            )

        current_row += sheet.max_row + 1  # Space between blocks

        # Add yellow separator row
        YELLOW_ROW_LENGTH = 50
        for col_idx in range(1, YELLOW_ROW_LENGTH + 1):
            combined_sheet.cell(row=current_row, column=col_idx).fill = yellow_fill
        current_row += 2

    # Apply test case columns blockwise
    row = 1
    while row <= combined_sheet.max_row:
        row_values = []
        for col in range(1, combined_sheet.max_column + 1):
            cell = combined_sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                row_values.append("")  # Treat merged (non-top-left) as empty
            else:
                value = str(cell.value).strip().lower() if cell.value else ""
                row_values.append(value)

        if "general" in row_values or "open" in row_values:
            # Determine where "relaxation years" is in the row
            try:
                start_col = row_values.index("relaxation years") + 2  # next column
            except ValueError:
                start_col = sheet.max_column + 1

            start_row = row + 1
            end_row = start_row

            while end_row <= combined_sheet.max_row:
                if all(combined_sheet.cell(row=end_row, column=col).value in [None, ""] for col in range(1, 6)):
                    break
                end_row += 1

            add_testcase_columns_func(combined_sheet, start_row - 1, end_row - 1, start_col)

            row = end_row
        else:
            row += 1

def append_testcase_other_details(sheet):
    last_col = get_last_data_column(sheet)
    # last_row = get_last_data_row(sheet)
    
    headers = ["EXPECTED OUTPUT", "DEFECT SEVERITY", "REMARKS"]
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, header in enumerate(headers):
        col_idx = last_col + i + 1
        col_letter = get_column_letter(col_idx)
        sheet[f"{col_letter}1"] = header
        sheet[f"{col_letter}1"].alignment = align

        if header == "EXPECTED OUTPUT":
            row = 2
            cell = sheet[f"{col_letter}{row}"]
            cell.value = "Pass"
            cell.fill = fill
            cell.alignment = align

    headers = ["EXPECTED OUTPUT", "DEFECT SEVERITY", "REMARKS"]
    fill = PatternFill(start_color="A6E22E", end_color="A6E22E", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

def process_excel(input_path, static_sheet_file, age_sheet_file, eligible_sheet_file):
    wb = openpyxl.load_workbook(input_path)

    delete_static_sheets(wb)
    insert_static_sheets(wb, static_sheet_file, 0)

    append_age_workbook_to_single_sheet(wb, age_sheet_file, add_testcase_columns_func)
    insert_static_sheets(wb, eligible_sheet_file, 1)

    for sheet in wb.worksheets:
        if sheet.title not in ["ReadMe", "Title","Age", "Other Details"]:
            add_testcase_columns(sheet)
        elif sheet.title == "Other Details":
            append_testcase_other_details(sheet)

    base, ext = os.path.splitext(input_path)
    output_path = f"{base}_testcases{ext}"
    wb.save(output_path)
    return output_path

# ---- Tkinter GUI ----
def browse_input():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        input_path.set(filepath)

def browse_static():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        static_path.set(filepath)

def browse_age():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        age_path.set(filepath)

def browse_eligible():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        eligible_path.set(filepath)

def run_processing():
    if not input_path.get() or not static_path.get() or not age_path.get() or not eligible_path.get():
        messagebox.showerror("Missing Input", "Please select all the files.")
        return
    try:
        output = process_excel(input_path.get(), static_path.get(), age_path.get(), eligible_path.get())
        messagebox.showinfo("Success", f"Test Case Document generated successfully!\n\nTest Case Document created at:\n{output}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n{str(e)}")

    # Reset fields
    input_path.set("")
    static_path.set("")
    age_path.set("")
    eligible_path.set("")

app = tk.Tk()
app.title("Unit Testcase Document Generator")
app.geometry("1000x1000")

input_path = tk.StringVar()
static_path = tk.StringVar()
age_path = tk.StringVar()
eligible_path = tk.StringVar()

tk.Label(app, text="Input SOW Excel File (SOW.xlsx)").pack(pady=5)
tk.Entry(app, textvariable=input_path, width=50).pack()
tk.Button(app, text="Browse", command=browse_input).pack(pady=5)

tk.Label(app, text="Static Sheets File (static_sheets.xlsx)").pack(pady=5)
tk.Entry(app, textvariable=static_path, width=50).pack()
tk.Button(app, text="Browse", command=browse_static).pack(pady=5)

tk.Label(app, text="Age Sheet (age.xlsx)").pack(pady=5)
tk.Entry(app, textvariable=age_path, width=50).pack()
tk.Button(app, text="Browse", command=browse_age).pack(pady=5)

tk.Label(app, text="Eligibility Sheet (eligibility.xlsx)").pack(pady=5)
tk.Entry(app, textvariable=eligible_path, width=50).pack()
tk.Button(app, text="Browse", command=browse_eligible).pack(pady=5)

tk.Button(app, text="Generate Test Case Document", command=run_processing, bg="green", fg="white").pack(pady=20)

app.mainloop()

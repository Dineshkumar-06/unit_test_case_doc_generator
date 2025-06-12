import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def is_row_empty(row):
    return all(cell.value is None for cell in row)

def get_last_data_column(sheet):
    max_col = 1
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value:
                max_col = max(max_col, i + 1)
    return max_col

def get_last_data_row(sheet):
    for row in reversed(list(sheet.iter_rows())):
        if not is_row_empty(row):
            return row[0].row
    return 1

def add_testcase_columns(sheet):
    last_col = get_last_data_column(sheet)
    last_row = get_last_data_row(sheet)

    headers = ["EXPECTED OUTPUT", "DEFECT SEVIARITY", "REMARKS"]
    fill = PatternFill(start_color="A6E22E", end_color="A6E22E", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center")

    for i, header in enumerate(headers):
        col_idx = last_col + i + 1
        col_letter = get_column_letter(col_idx)
        sheet[f"{col_letter}1"] = header
        sheet[f"{col_letter}1"].alignment = align

        if header == "EXPECTED OUTPUT":
            for row in range(2, last_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.value = "Pass"
                cell.fill = fill
                cell.alignment = align

def delete_static_sheets(wb):
    for name in ["ReadMe", "Title"]:
        if name in wb.sheetnames:
            std = wb[name]
            wb.remove(std)

# def insert_static_sheets(wb_target, static_sheet_file):
#     wb_static = openpyxl.load_workbook(static_sheet_file)

#     for name in reversed(wb_static.sheetnames):
#         sheet = wb_static[name]
#         new_sheet = wb_target.copy_worksheet(sheet)
#         new_sheet.title = name
#         wb_target._sheets.insert(0, wb_target._sheets.pop())  # bring to front

import copy
from openpyxl.cell.cell import Cell  # Import Cell class to check type

def insert_static_sheets(wb_target, static_sheet_file):
    wb_static = openpyxl.load_workbook(static_sheet_file)

    for sheet in reversed(wb_static.sheetnames):
        source = wb_static[sheet]
        new_sheet = wb_target.create_sheet(title=sheet, index=0)

        for row in source.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell):
                    continue  # Skip MergedCell or other non-standard cells

                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                if cell.has_style:
                    if cell.font:
                        new_cell.font = copy.copy(cell.font)
                    if cell.border:
                        new_cell.border = copy.copy(cell.border)
                    if cell.fill:
                        new_cell.fill = copy.copy(cell.fill)
                    if cell.number_format:
                        new_cell.number_format = copy.copy(cell.number_format)
                    if cell.protection:
                        new_cell.protection = copy.copy(cell.protection)
                    if cell.alignment:
                        new_cell.alignment = copy.copy(cell.alignment)

        # Copy merged cell ranges
        for merged_range in source.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))

        # Copy column widths
        for col in source.column_dimensions:
            new_sheet.column_dimensions[col].width = source.column_dimensions[col].width

        # Copy row heights
        for row_idx in source.row_dimensions:
            new_sheet.row_dimensions[row_idx].height = source.row_dimensions[row_idx].height

def process_excel(input_path, static_sheet_file):
    wb = openpyxl.load_workbook(input_path)

    delete_static_sheets(wb)
    insert_static_sheets(wb, static_sheet_file)

    for sheet in wb.worksheets:
        if sheet.title not in ["ReadMe", "Title"]:
            add_testcase_columns(sheet)

    base, ext = os.path.splitext(input_path)
    output_path = f"{base}_testcases{ext}"
    wb.save(output_path)
    return output_path

# ---- Tkinter GUI ----
def browse_input():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        input_path.set(filepath)

def browse_static():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        static_path.set(filepath)

def run_processing():
    if not input_path.get() or not static_path.get():
        messagebox.showerror("Missing Input", "Please select both input and static files.")
        return
    try:
        output = process_excel(input_path.get(), static_path.get())
        messagebox.showinfo("Success", f"✅ Test case Excel created:\n{output}")
    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n{str(e)}")

app = tk.Tk()
app.title("Test Case Excel Generator")
app.geometry("500x250")

input_path = tk.StringVar()
static_path = tk.StringVar()

tk.Label(app, text="Input Excel File (SOW.xlsx)").pack(pady=5)
tk.Entry(app, textvariable=input_path, width=50).pack()
tk.Button(app, text="Browse", command=browse_input).pack(pady=5)

tk.Label(app, text="Static Sheets File (StaticSheets.xlsx)").pack(pady=5)
tk.Entry(app, textvariable=static_path, width=50).pack()
tk.Button(app, text="Browse", command=browse_static).pack(pady=5)

tk.Button(app, text="Generate Test Case Document", command=run_processing, bg="green", fg="white").pack(pady=20)

app.mainloop()
